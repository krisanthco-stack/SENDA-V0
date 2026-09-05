from pathlib import Path
from app.repository import Repository
from app.services.exports import export_json, export_csv, export_xlsx

def test_exports_do_not_contain_enlace_column(tmp_path):
    repo=Repository(tmp_path/'db.sqlite')
    repo.insert_movements([{'folio':'4-281400','finca':'4-281400','fecha':'2026-06-01','fuente':'FINCAS','operacion':'COMPRAVENTA','enlace':'https://example.test/a'}],1)
    jp=export_json(repo,{},tmp_path/'out.json')
    cp=export_csv(repo,{},tmp_path/'out.csv')
    xp=export_xlsx(repo,{},tmp_path/'out.xlsx')
    assert 'enlace' not in jp.read_text('utf-8').lower()
    assert 'enlace' not in cp.read_text('utf-8-sig').splitlines()[0].lower()
    from openpyxl import load_workbook
    wb=load_workbook(xp,read_only=True,data_only=True)
    try:
        headers=[str(x or '').lower() for x in next(wb['Movimientos'].iter_rows(values_only=True))]
        assert 'enlace' not in headers
    finally:
        wb.close()
