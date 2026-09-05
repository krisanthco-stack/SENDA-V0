import json, tempfile, unittest, zipfile
from pathlib import Path
from app.repository import Repository
from app.importers.engine import ImportEngine, detect_format, detect_source

class ImporterTests(unittest.TestCase):
    def setUp(self):
        self.td=tempfile.TemporaryDirectory(); self.root=Path(self.td.name); self.repo=Repository(self.root/'db.sqlite'); self.engine=ImportEngine(self.repo)
    def tearDown(self): self.td.cleanup()

    def test_detects_content_not_only_extension(self):
        p=self.root/'fake.xls'; p.write_text('PROVINCIA;NUMERO;DERECHO\n4;123;1\n', encoding='utf-8')
        self.assertEqual(detect_format(p),'delimited')
        self.assertEqual(detect_source(p, ['PROVINCIA','NUMERO','DERECHO']),'FINCAS')

    def test_utf8_bom_does_not_corrupt_first_header(self):
        p=self.root/'convertido.csv'
        p.write_text('PROVINCIA;NUMERO;DERECHO\n4;200103;1\n',encoding='utf-8-sig')
        from app.importers.engine import iter_delimited
        row=next(iter_delimited(p))
        self.assertIn('PROVINCIA',row)
        self.assertEqual(row['PROVINCIA'],'4')

    def test_catalog_rules_and_operation_class_are_applied(self):
        (self.root/'CATALOGO_COD_OPERACIONES.TXT').write_text('"PE";1;"COMPRAVENTA"\n',encoding='utf-8')
        (self.root/'CATALOGO_COD_DERECHOS.TXT').write_text('"D";"DOMINIO"\n',encoding='utf-8')
        (self.root/'CATALOGO_COD_STATUS.txt').write_text('VALOR D = CERRADA\nVALOR NULL = ACTIVAS\nVALOR B y EN BLANCO SE DEBEN OMITIR\n',encoding='utf-8')
        (self.root/'CATALOGO_COD_CLASE_RESP.txt').write_text('9; VALOR OMITIR\nC; EN CONTRA\n',encoding='utf-8')
        csvp=self.root/'Fincas.csv'
        csvp.write_text('PROVINCIA;NUMERO;DERECHO;COD_DERECHO;COD_OPERACION;CLASE_CODIGO;FECHA_ULT_ACT;STATUS;CLASE_RESP\n4;200103;1;D;PE;1;01/03/2026;NULL;C\n4;200104;1;D;PE;1;01/03/2026;B;C\n4;200105;1;D;PE;1;01/03/2026;D;9\n',encoding='utf-8')
        result=self.engine.import_paths([*self.root.glob('CATALOGO*'),csvp], year=2026, quarter='T1', district='Horquetas')
        self.assertEqual(result['inserted'],1)
        row=self.repo.list_movements({'district':'HORQUETAS'})[0]
        self.assertEqual(row['codigo'],'PE1'); self.assertEqual(row['operacion'],'COMPRAVENTA'); self.assertEqual(row['derecho'],'DOMINIO')
        self.assertEqual(row['trimestre'],'T1'); self.assertEqual(row['mes'],3)

    def test_real_biff_xls_fixture_is_read(self):
        fixture=Path(__file__).parent/'fixtures'/'Fincas_real.xls'
        result=self.engine.import_paths([fixture], year=2026, quarter='T2', district='Puerto Viejo')
        self.assertEqual(result['inserted'],1)
        row=self.repo.list_movements({'district':'PUERTO VIEJO'})[0]
        self.assertEqual(row['folio'],'4-200103-001')
        self.assertEqual(row['fecha'],'2026-06-01')

    def test_json_array_and_zip_are_imported(self):
        jp=self.root/'movimientos.json'; jp.write_text(json.dumps([{'folio':'4-300-001','fecha':'2026-07-04','fuente':'HISTORICOS','codigo':'X'}]),encoding='utf-8')
        inside='PROVINCIA;NUMERO;DERECHO;FECHA_ULT_ACT\n4;400;1;01/08/2026\n'
        zp=self.root/'SARAPIQUI.zip'
        with zipfile.ZipFile(zp,'w',compression=zipfile.ZIP_DEFLATED) as z:z.writestr('Fincas_SARAPIQUI.csv',inside)
        result=self.engine.import_paths([jp,zp], year=2026, quarter='T3', district='La Virgen')
        self.assertEqual(result['inserted'],2)
        self.assertEqual(len(self.repo.list_movements({'quarter':'T3','district':'LA VIRGEN'})),2)



    def test_excel_url_columns_are_not_imported_as_movement_links(self):
        p=self.root/'Fincas_con_enlace.csv'
        p.write_text('PROVINCIA;NUMERO;DERECHO;FECHA_ULT_ACT;ENLACE;URL\n4;281400;0;01/06/2026;https://example.test/a;https://example.test/b\n',encoding='utf-8')
        result=self.engine.import_paths([p],year=2026,quarter='T2',district='Horquetas')
        self.assertEqual(result['inserted'],1)
        row=self.repo.list_movements({'search':'281400'})[0]
        self.assertNotIn('enlace',row)

    def test_zip_rejects_prefix_path_traversal(self):
        archive=self.root/'malicioso.zip'
        target=self.root/'extract'
        target.mkdir()
        with zipfile.ZipFile(archive,'w',compression=zipfile.ZIP_DEFLATED) as z:
            z.writestr('../extract_evil/Fincas.csv','PROVINCIA;NUMERO;DERECHO\n4;999;1\n')
        from app.importers.engine import safe_extract_zip
        with self.assertRaises(RuntimeError):
            safe_extract_zip(archive,target)
        self.assertFalse((self.root/'extract_evil'/'Fincas.csv').exists())

    def test_nested_zip_with_catalog_and_data_is_imported(self):
        inner=self.root/'inner.zip'
        with zipfile.ZipFile(inner,'w',compression=zipfile.ZIP_DEFLATED) as z:
            z.writestr('CATALOGO_COD_OPERACIONES.TXT','"PE";1;"COMPRAVENTA"\n')
            z.writestr('datos/Fincas.csv','PROVINCIA;NUMERO;DERECHO;COD_OPERACION;CLASE_CODIGO;FECHA_ULT_ACT\n4;777;1;PE;1;15/03/2026\n')
        outer=self.root/'SARAPIQUI_T1.zip'
        with zipfile.ZipFile(outer,'w',compression=zipfile.ZIP_DEFLATED) as z:
            z.write(inner,'paquetes/inner.zip')
        result=self.engine.import_paths([outer],year=2026,quarter='T1',district='Horquetas')
        self.assertEqual(result['inserted'],1)
        row=self.repo.list_movements({'district':'HORQUETAS'})[0]
        self.assertEqual(row['operacion'],'COMPRAVENTA')
        self.assertEqual(row['archivo_origen'],'Fincas.csv')


    def test_zip_folder_assigns_district_per_file(self):
        zp=self.root/'SARAPIQUI.zip'
        with zipfile.ZipFile(zp,'w',compression=zipfile.ZIP_DEFLATED) as z:
            z.writestr('PUERTO VIEJO/Fincas.csv','PROVINCIA;NUMERO;DERECHO;FECHA_ULT_ACT\n4;501;1;01/03/2026\n')
            z.writestr('LAS HORQUETAS/Fincas.csv','PROVINCIA;NUMERO;DERECHO;FECHA_ULT_ACT\n4;502;1;01/03/2026\n')
        result=self.engine.import_paths([zp],year=2026,quarter='T1',district='')
        self.assertEqual(result['inserted'],2)
        self.assertEqual(len(self.repo.list_movements({'district':'PUERTO VIEJO'})),1)
        self.assertEqual(len(self.repo.list_movements({'district':'HORQUETAS'})),1)

    def test_import_pipeline_is_lazy_not_eager(self):
        import app.importers.engine as eng
        p=self.root/'Fincas.csv'; p.write_text('x',encoding='utf-8')
        original=eng.iter_rows
        def rows(_):
            yield {'PROVINCIA':'4','NUMERO':'1','DERECHO':'1','FECHA':'2026-03-01'}
            raise AssertionError('pipeline consumed beyond repository demand')
        eng.iter_rows=rows
        original_insert=self.repo.insert_movements
        self.repo.insert_movements=lambda iterable, import_id: (next(iter(iterable)) and 1)
        try:
            result=self.engine.import_paths([p],year=2026,quarter='T1',district='Horquetas')
            self.assertEqual(result['inserted'],1)
        finally:
            eng.iter_rows=original; self.repo.insert_movements=original_insert

    def test_reimporting_same_file_hash_does_not_duplicate(self):
        p=self.root/'Fincas.csv'
        p.write_text('PROVINCIA;NUMERO;DERECHO;FECHA_ULT_ACT\n4;901;1;01/03/2026\n',encoding='utf-8')
        first=self.engine.import_paths([p],year=2026,quarter='T1',district='Horquetas')
        second=self.engine.import_paths([p],year=2026,quarter='T1',district='Horquetas')
        self.assertEqual(first['inserted'],1)
        self.assertEqual(second['inserted'],0)
        self.assertTrue(second.get('duplicate_import'))
        self.assertEqual(len(self.repo.list_movements({'district':'HORQUETAS'})),1)

    def test_shared_exact_source_row_is_deduplicated_across_different_files(self):
        a=self.root/'Fincas_A.csv'
        b=self.root/'Fincas_B.csv'
        header='PROVINCIA;NUMERO;DERECHO;FECHA_ULT_ACT;NOTA\n'
        a.write_text(header+'4;902;1;01/03/2026;MISMA\n',encoding='utf-8')
        b.write_text(header+'4;902;1;01/03/2026;MISMA\n4;903;1;01/03/2026;NUEVA\n',encoding='utf-8')
        first=self.engine.import_paths([a],year=2026,quarter='T1',district='Horquetas')
        second=self.engine.import_paths([b],year=2026,quarter='T1',district='Horquetas')
        self.assertEqual(first['inserted'],1)
        self.assertEqual(second['inserted'],1)
        self.assertGreaterEqual(second.get('duplicates',0),1)
        self.assertEqual(len(self.repo.list_movements({'district':'HORQUETAS'})),2)

    def test_same_logical_movement_with_auxiliary_metadata_change_is_deduplicated(self):
        a=self.root/'Fincas_A.csv'
        b=self.root/'Fincas_B.csv'
        header='PROVINCIA;NUMERO;DERECHO;FECHA_ULT_ACT;NOTA\n'
        a.write_text(header+'4;905;1;01/03/2026;EXPORTACION A\n',encoding='utf-8')
        b.write_text(header+'4;905;1;01/03/2026;EXPORTACION B\n',encoding='utf-8')
        first=self.engine.import_paths([a],year=2026,quarter='T1',district='Horquetas')
        second=self.engine.import_paths([b],year=2026,quarter='T1',district='Horquetas')
        self.assertEqual(first['inserted'],1)
        self.assertEqual(second['inserted'],0)
        self.assertEqual(second.get('duplicates',0),1)
        self.assertEqual(len(self.repo.list_movements({'district':'HORQUETAS'})),1)

    def test_distinct_logical_movements_are_not_collapsed_by_dedupe(self):
        a=self.root/'Fincas_A.csv'
        b=self.root/'Fincas_B.csv'
        a.write_text('PROVINCIA;NUMERO;DERECHO;FECHA_ULT_ACT;COD_OPERACION\n4;904;1;01/03/2026;PE1\n',encoding='utf-8')
        b.write_text('PROVINCIA;NUMERO;DERECHO;FECHA_ULT_ACT;COD_OPERACION\n4;904;1;01/03/2026;PG1\n',encoding='utf-8')
        first=self.engine.import_paths([a],year=2026,quarter='T1',district='Horquetas')
        second=self.engine.import_paths([b],year=2026,quarter='T1',district='Horquetas')
        self.assertEqual(first['inserted'],1)
        self.assertEqual(second['inserted'],1)
        self.assertEqual(second.get('duplicates',0),0)
        self.assertEqual(len(self.repo.list_movements({'district':'HORQUETAS'})),2)

    def test_xlsx_is_imported_without_excel_installed(self):
        from openpyxl import Workbook
        p=self.root/'Fincas.xlsx'
        wb=Workbook(); ws=wb.active
        ws.append(['PROVINCIA','NUMERO','DERECHO','FECHA_ULT_ACT'])
        ws.append([4,903,1,'2026-03-01'])
        wb.save(p); wb.close()
        result=self.engine.import_paths([p],year=2026,quarter='T1',district='Horquetas')
        self.assertEqual(result['inserted'],1)
        self.assertEqual(self.repo.list_movements({'district':'HORQUETAS'})[0]['folio'],'4-903-001')

    def test_numero_is_authoritative_when_folio_conflicts(self):
        from app.importers.engine import normalize_row, catalog_context
        ctx=catalog_context(self.repo)
        row=normalize_row(
            {'PROVINCIA':'4','NUMERO':'140302','DERECHO':'1','FOLIO':'4-999999-001','FECHA_ULT_ACT':'01/03/2026'},
            source='FINCAS',year=2026,quarter='T1',district='Horquetas',source_file='Fincas.xls',ctx=ctx)
        self.assertEqual(row['folio'],'4-140302-001')

    def test_numero_without_derecho_is_preserved_as_finca(self):
        from app.importers.engine import normalize_row, catalog_context
        ctx=catalog_context(self.repo)
        row=normalize_row(
            {'PROVINCIA':'4','NUMERO':'140302','FECHA_ULT_ACT':'01/03/2026'},
            source='FINCAS',year=2026,quarter='T1',district='Horquetas',source_file='Fincas.xls',ctx=ctx)
        self.assertEqual(row['folio'],'4-140302')

    def test_numero_with_zero_derecho_does_not_become_blank_or_fake_000(self):
        from app.importers.engine import normalize_row, catalog_context
        ctx=catalog_context(self.repo)
        row=normalize_row(
            {'PROVINCIA':'4','NUMERO':'141174','DERECHO':'0','FECHA_ULT_ACT':'01/03/2026'},
            source='HISTORICOS',year=2026,quarter='T1',district='Horquetas',source_file='Historicos.xls',ctx=ctx)
        self.assertEqual(row['folio'],'4-141174')
        self.assertNotIn('-000',row['folio'])

    def test_numero_is_preserved_when_province_is_missing(self):
        from app.importers.engine import normalize_row, catalog_context
        ctx=catalog_context(self.repo)
        row=normalize_row(
            {'NUMERO':'140302','FECHA_ULT_ACT':'01/03/2026'},
            source='FINCAS',year=2026,quarter='T1',district='Horquetas',source_file='Fincas.xls',ctx=ctx)
        self.assertEqual(row['folio'],'140302')

    def test_numero_de_finca_header_alias_is_recognized(self):
        from app.importers.engine import normalize_row, catalog_context
        ctx=catalog_context(self.repo)
        row=normalize_row(
            {'PROVINCIA':'HEREDIA','NUMERO_DE_FINCA':'140302','FECHA_ULT_ACT':'01/03/2026'},
            source='FINCAS',year=2026,quarter='T1',district='Horquetas',source_file='Fincas.xlsx',ctx=ctx)
        self.assertEqual(row['folio'],'4-140302')


    def test_real_export_empty_status_is_active_and_numero_zero_is_folio(self):
        p=self.root/'Fincas_SARAPIQUI.xls'
        p.write_text(
            'PROVINCIA\tCANTON\tDISTRITO\tNUMERO\tDERECHO\tCOD_DERECHO\tFECHA_ULT_ACT\tSTATUS\n'
            '4\t10\t3\t140302\t0\tD\t2026-03-02 08:00:00.0000000\t\n',
            encoding='utf-8'
        )
        result=self.engine.import_paths([p],year=2026,quarter='T1',district='')
        self.assertEqual(result['inserted'],1)
        self.assertEqual(result['skipped'],0)
        row=self.repo.list_movements({'quarter':'T1'})[0]
        self.assertEqual(row['folio'],'4-140302')

    def test_real_export_whitespace_status_is_blank_and_omitted(self):
        p=self.root/'Fincas_SARAPIQUI.xls'
        p.write_text(
            'PROVINCIA\tNUMERO\tDERECHO\tFECHA_ULT_ACT\tSTATUS\n'
            '4\t141174\t0\t2026-03-02 08:00:00.0000000\t \n',
            encoding='utf-8'
        )
        result=self.engine.import_paths([p],year=2026,quarter='T1',district='')
        self.assertEqual(result['inserted'],0)
        self.assertEqual(result['skipped'],1)

    def test_cerrada_with_numero_and_no_derecho_keeps_base_folio(self):
        p=self.root/'Cerradas_SARAPIQUI.xls'
        p.write_text(
            'PROVINCIA\tCANTON\tDISTRITO\tNUMERO\tFECHA_ULT_ACT\tSTATUS\n'
            '4\t10\t3\t134987\t2025-12-03 00:00:00.0000000\tD\n',
            encoding='utf-8'
        )
        result=self.engine.import_paths([p],year=2026,quarter='T1',district='')
        self.assertEqual(result['inserted'],1)
        row=self.repo.list_movements({})[0]
        self.assertEqual(row['folio'],'4-134987')
        self.assertEqual(row['categoria'],'CERRADAS')

    def test_supported_desktop_import_contract_lists_all_required_formats(self):
        from app.importers.engine import SUPPORTED_IMPORT_FORMATS
        self.assertEqual(set(SUPPORTED_IMPORT_FORMATS), {'XLS','XLSX','CSV','JSON','TXT','ZIP','RAR'})

if __name__=='__main__': unittest.main()
