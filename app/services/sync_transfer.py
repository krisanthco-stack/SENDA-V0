from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path

from .exports import COLS, iter_pages

FORMAT_NAME = 'SENDA_TRANSFER'
FORMAT_VERSION = 2
CASE_COLS = [
    'id','folio','plano','distrito','status','responsable','prioridad','note',
    'control_started_at','finalized_at','management_started_at','management_state','created_at','updated_at'
]
AUDIT_COLS = [
    'folio','plano','action','previous_status','new_status','note','payload_json','created_at'
]


def _now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def detect_sync_file(path: str | Path) -> bool:
    path = Path(path)
    if path.suffix.lower() == '.json':
        try:
            with path.open('r', encoding='utf-8-sig', errors='replace') as f:
                prefix = f.read(32768)
            return '"SENDA_TRANSFER"' in prefix and '"formato"' in prefix
        except Exception:
            return False
    if path.suffix.lower() == '.xlsx':
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                if 'RESUMEN' not in wb.sheetnames:
                    return False
                values = {}
                for row in wb['RESUMEN'].iter_rows(min_row=1, max_row=12, values_only=True):
                    if row and row[0] is not None:
                        values[str(row[0]).strip().lower()] = '' if len(row) < 2 or row[1] is None else str(row[1]).strip()
                return values.get('formato') == FORMAT_NAME
            finally:
                wb.close()
        except Exception:
            return False
    return False


def export_sync_json(repo, target: str | Path) -> Path:
    target = Path(target)
    cases = repo.export_case_rows()
    audits = repo.export_audit_rows()
    with target.open('w', encoding='utf-8') as f:
        f.write('{')
        # Explicit keys keep the format recognizable from the file prefix.
        f.write('"sistema":"SENDA.V0",')
        f.write('"formato":"SENDA_TRANSFER",')
        f.write(f'"version_formato":{FORMAT_VERSION},')
        f.write('"exported_at":'); json.dump(_now_iso(), f, ensure_ascii=False); f.write(',')
        f.write('"movimientos":[')
        first = True
        for page in iter_pages(repo, {}, page_size=5000):
            for row in page:
                if not first:
                    f.write(',')
                json.dump({k: row.get(k, '') for k in COLS}, f, ensure_ascii=False, default=str)
                first = False
        f.write('],"expedientes":')
        json.dump([{k: r.get(k, '') for k in CASE_COLS} for r in cases], f, ensure_ascii=False, default=str)
        f.write(',"auditoria":')
        json.dump([{k: r.get(k, '') for k in AUDIT_COLS} for r in audits], f, ensure_ascii=False, default=str)
        f.write(',"selecciones":')
        json.dump(repo.export_selection_rows(), f, ensure_ascii=False, default=str)
        f.write(',"referencias_registrales":')
        json.dump(repo.export_registry_reference_rows(), f, ensure_ascii=False, default=str)
        f.write(',"cedulas_juridicas":')
        json.dump(repo.export_legal_entity_rows(), f, ensure_ascii=False, default=str)
        f.write(',"representantes":')
        json.dump(repo.export_legal_representative_rows(), f, ensure_ascii=False, default=str)
        f.write('}')
    return target


def export_sync_xlsx(repo, target: str | Path) -> Path:
    target = Path(target)
    import xlsxwriter
    wb = xlsxwriter.Workbook(str(target), {'constant_memory': True})
    header = wb.add_format({'bold': True, 'bg_color': '#DCE6F1', 'border': 1})
    subheader = wb.add_format({'bold': True, 'bg_color': '#EEF4F9'})

    summary = wb.add_worksheet('RESUMEN')
    for r, (key, value) in enumerate([
        ('SISTEMA', 'SENDA.V0'), ('FORMATO', FORMAT_NAME), ('VERSION_FORMATO', FORMAT_VERSION),
        ('EXPORTADO_EN', _now_iso()), ('DESCRIPCION', 'Base SENDA fusionable entre computadoras')
    ]):
        summary.write(r, 0, key, subheader); summary.write(r, 1, value)
    summary.set_column(0, 0, 24); summary.set_column(1, 1, 58)

    mov = wb.add_worksheet('MOVIMIENTOS')
    for c, name in enumerate(COLS): mov.write(0, c, name.upper(), header)
    rix = 1
    for page in iter_pages(repo, {}, page_size=5000):
        for row in page:
            for c, name in enumerate(COLS): mov.write(rix, c, row.get(name, ''))
            rix += 1
    mov.freeze_panes(1, 0); mov.autofilter(0, 0, max(0, rix-1), len(COLS)-1); mov.set_column(0, len(COLS)-1, 16)
    mov.set_column(COLS.index('operacion'), COLS.index('operacion'), 36)

    cases = wb.add_worksheet('EXPEDIENTES')
    for c, name in enumerate(CASE_COLS): cases.write(0, c, name.upper(), header)
    for rix, row in enumerate(repo.export_case_rows(), start=1):
        for c, name in enumerate(CASE_COLS): cases.write(rix, c, row.get(name, ''))
    cases.freeze_panes(1, 0); cases.set_column(0, len(CASE_COLS)-1, 18); cases.set_column(CASE_COLS.index('note'), CASE_COLS.index('note'), 42)

    audits = wb.add_worksheet('AUDITORIA')
    for c, name in enumerate(AUDIT_COLS): audits.write(0, c, name.upper(), header)
    for rix, row in enumerate(repo.export_audit_rows(), start=1):
        for c, name in enumerate(AUDIT_COLS): audits.write(rix, c, row.get(name, ''))
    audits.freeze_panes(1, 0); audits.set_column(0, len(AUDIT_COLS)-1, 20); audits.set_column(AUDIT_COLS.index('payload_json'), AUDIT_COLS.index('payload_json'), 42)

    selections = wb.add_worksheet('SELECCIONES')
    sel_cols=['folio','plano','signature']
    for c,name in enumerate(sel_cols):selections.write(0,c,name.upper(),header)
    for rix,row in enumerate(repo.export_selection_rows(),start=1):
        for c,name in enumerate(sel_cols):selections.write(rix,c,row.get(name,''))
    selections.freeze_panes(1,0);selections.set_column(0,1,20);selections.set_column(2,2,70)

    refs = wb.add_worksheet('REFERENCIAS')
    ref_cols=['kind','district','finca','plan_key','source_file','raw_json']
    for c,name in enumerate(ref_cols):refs.write(0,c,name.upper(),header)
    for rix,row in enumerate(repo.export_registry_reference_rows(),start=1):
        for c,name in enumerate(ref_cols):refs.write(rix,c,row.get(name,''))
    refs.freeze_panes(1,0);refs.set_column(0,4,20);refs.set_column(5,5,80)

    leg = wb.add_worksheet('CED_JURIDICAS')
    leg_cols=['cedula_juridica','razon_social','domicilio','fecha_inicio','fecha_vence','source_file']
    for c,name in enumerate(leg_cols):leg.write(0,c,name.upper(),header)
    for rix,row in enumerate(repo.export_legal_entity_rows(),start=1):
        for c,name in enumerate(leg_cols):leg.write(rix,c,row.get(name,''))
    leg.freeze_panes(1,0);leg.set_column(0,5,24)

    reps = wb.add_worksheet('REPRESENTANTES')
    rep_cols=['cedula_juridica','numero_identificacion','nombre','apellido_1','apellido_2','cargo','representacion','source_file']
    for c,name in enumerate(rep_cols):reps.write(0,c,name.upper(),header)
    for rix,row in enumerate(repo.export_legal_representative_rows(),start=1):
        for c,name in enumerate(rep_cols):reps.write(rix,c,row.get(name,''))
    reps.freeze_panes(1,0);reps.set_column(0,7,22)

    wb.close()
    return target


def _read_sheet(ws) -> list[dict]:
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [str(v or '').strip().lower() for v in next(rows)]
    except StopIteration:
        return []
    out = []
    for values in rows:
        row = {h: values[i] for i, h in enumerate(headers) if h and i < len(values)}
        if any(v not in (None, '') for v in row.values()):
            out.append(row)
    return out


def _load_sync_payload(path: Path) -> dict:
    if path.suffix.lower() == '.json':
        with path.open('r', encoding='utf-8-sig') as f:
            payload = json.load(f)
        if payload.get('formato') != FORMAT_NAME:
            raise ValueError('El JSON no es una base SENDA fusionable.')
        return payload
    if path.suffix.lower() == '.xlsx':
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            if not {'RESUMEN', 'MOVIMIENTOS', 'EXPEDIENTES', 'AUDITORIA'} <= set(wb.sheetnames):
                raise ValueError('El Excel no contiene todas las hojas de una base SENDA fusionable.')
            meta = {}
            for row in wb['RESUMEN'].iter_rows(values_only=True):
                if row and row[0] is not None:
                    meta[str(row[0]).strip().lower()] = row[1] if len(row) > 1 else ''
            if str(meta.get('formato', '')) != FORMAT_NAME:
                raise ValueError('El Excel no es una base SENDA fusionable.')
            return {
                'sistema': meta.get('sistema', 'SENDA.V0'), 'formato': FORMAT_NAME,
                'version_formato': int(meta.get('version_formato') or FORMAT_VERSION),
                'exported_at': meta.get('exportado_en', ''),
                'movimientos': _read_sheet(wb['MOVIMIENTOS']),
                'expedientes': _read_sheet(wb['EXPEDIENTES']),
                'auditoria': _read_sheet(wb['AUDITORIA']),
                'selecciones': _read_sheet(wb['SELECCIONES']) if 'SELECCIONES' in wb.sheetnames else [],
                'referencias_registrales': _read_sheet(wb['REFERENCIAS']) if 'REFERENCIAS' in wb.sheetnames else [],
                'cedulas_juridicas': _read_sheet(wb['CED_JURIDICAS']) if 'CED_JURIDICAS' in wb.sheetnames else [],
                'representantes': _read_sheet(wb['REPRESENTANTES']) if 'REPRESENTANTES' in wb.sheetnames else [],
            }
        finally:
            wb.close()
    raise ValueError('Formato de base SENDA no soportado.')


def import_sync_file(repo, path: str | Path) -> dict:
    path = Path(path)
    if not detect_sync_file(path):
        raise ValueError('El archivo no es una base SENDA fusionable.')
    payload = _load_sync_payload(path)
    result = repo.merge_sync_payload(payload, source_name=path.name)
    result['sync'] = True
    result['source'] = path.name
    return result
